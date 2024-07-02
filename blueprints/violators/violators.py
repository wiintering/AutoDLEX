from flask import Blueprint, redirect,render_template,request,url_for,make_response,flash
from db_utils import get_db_connection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import os





db_connection = get_db_connection()
cursor = db_connection.cursor()

violator_bp = Blueprint("violators", __name__, template_folder="templates")

@violator_bp.route('/violator_list')
def violator_list():
    cursor.execute("SELECT * FROM violators_data WHERE status = 'Unsettled'")
    data =  cursor.fetchall()
    return render_template('Violator_list.html',violators = data)

@violator_bp.route('/view_violator')
def view_violator():
    violator_id = request.args.get('violator_id')
    query = "SELECT * FROM violators_data WHERE violators_id = %s"
    cursor.execute(query, (violator_id,))
    violator_data = cursor.fetchone()
    if violator_data:
        return render_template('violators_data/view-violators-data.html', violator_data=violator_data)
    else:
        return "Violator not found", 404
    
    
#viewing of violators data specific 
@violator_bp.route('/view_violators_data')
def view_violators_data():
    violator_id = request.args.get('violator_id')
    query = "SELECT * FROM violators_data WHERE violators_id  = %s"
    cursor.execute(query, (violator_id,))
    violator_data = cursor.fetchone()
    
    #getting the tct number of the violator
    tct_number = " Select tct_number from violators_data where violators_id = %s"
    cursor.execute(tct_number, (violator_id,))
    tct_number = cursor.fetchone()
    
    tct_number = tct_number[0]
    
    extracted_data = "Select * from extracted_data where tct_number = %s"
    cursor.execute(extracted_data, (tct_number,))
    extracted_data = cursor.fetchone()
    
   
    
    return render_template('violators_data/view-violators-data.html',extracted_data = extracted_data , violator_data = violator_data)

    
# Editing/Updating  violators data 
from flask import render_template, request

@violator_bp.route('/edit_violators_data', methods=['GET', 'POST'])  
def edit_violators_data():
    msg = ''
    violator_id = request.args.get('violator_id')
    # Assuming you have a function to get violator data from the database
    query = "SELECT * FROM violators_data WHERE violators_id  = %s"
    cursor.execute(query, (violator_id,))
    violator_data = cursor.fetchone()
    
    if request.method == 'POST':
        violator_id = request.form['violators_id']
        violation = request.form['violation']
        time = request.form['time']
        date = request.form['date']
        barangay = request.form['barangay']
        plateNumber = request.form['plateNumber']
        vehicle = request.form['vehicle']
        status = request.form['status']
       
        
        cursor.execute("UPDATE violators_data SET violation=%s, time=%s, date=%s, barangay=%s, plateNumber=%s, vehicle=%s ,status=%s WHERE violators_id  = %s",
                       (violation, time, date, barangay, plateNumber, vehicle, status,violator_id))
        db_connection.commit() 
        flash('Violator data updated successfully!', 'success')
        return redirect(url_for('violators.violator_list'))

    # Assuming violator_id is passed as a query parameter

   
    return render_template('violators_data/edit-violators-data.html', violator_data=violator_data)


        
#deleting of violators data        
@violator_bp.route('/delete_violator')
def delete_violator():
    violators_id = request.args.get('violator_id').split(',')
    for violator in violators_id:
        cursor.execute("DELETE FROM violators_data WHERE violators_id = %s", (violator,))
        db_connection.commit() 
    return redirect(url_for('settled_reports'))






@violator_bp.route('/excellReports')
def excellReports():
    violator_ids = request.args.get('violator_ids').split(',')
    all_data = []
    
   
    # Connect to the database (already initialized db_connection and cursor assumed)

    for violator_id in violator_ids:
            # Fetch tct_number from violators_data table for the current violator_id
            cursor.execute("SELECT tct_number FROM violators_data WHERE violators_id = %s", (violator_id,))
            ticket_numbers = cursor.fetchall()

            for ticket_number_tuple in ticket_numbers:
                ticket_number = ticket_number_tuple[0]  # Access the first element (tct_number) of the tuple

                # Fetch data from settled_report using ticket_number
                cursor.execute("SELECT * FROM reports WHERE tct_number = %s", (ticket_number,))
                report_data = cursor.fetchall()

                # Extend all_data with the fetched report_data
                all_data.extend(report_data)
        

    # Create a new workbook
    wb = Workbook()
    sheet = wb.active

    # Set up the margins
    top_margin = 6
    left_margin = 2  # At least one cell from the left

    # Set up the logo image paths
  


    # Set up the text for the headers
    headers = [
        "Republika ng Pilipinas",
        "Lungsod ng Santa Rosa",
        "Lalawigan ng Laguna",
        "SANGAY NG PAMAMAHALA AT PAGPAPATUPAD NG TRAPIKO",
        "(CITY TRAFFIC MANAGEMENT AND ENFORCEMENT UNIT)"
    ]

    # Add the text with the required margins and background colors
    for row_index, header in enumerate(headers, start=top_margin):
        cell = sheet.cell(row=row_index, column=left_margin)
        cell.value = header
        cell.font = Font(size=12, bold=True)
        cell.alignment = Alignment(horizontal='center')
        if row_index % 2 == 0:
            # Set light gray background for even rows (headers)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        else:
            # No background for odd rows
            pass
        sheet.merge_cells(start_row=row_index, start_column=left_margin, end_row=row_index, end_column=left_margin + 9)

    # Set up the header for the table
    table_header_row = top_margin + len(headers) + 1
    table_headers = [
        "Report_ID",
        "TRAFFIC CITATION TICKET NO.",
        "NAME OF APPREHENDED DRIVER/OPERATOR",
        "VIOLATION(S)",
        "PLACE OF APPREHENSION",
        "DRIVER'S LICENSE NO.",
        " PLATE NO.",
        "Vehicle",
        "DATE/TIME OF APPREHENSION",
        "RECEIVED BY (NAME AND SIGNATURE)"
    ]

    for col_index, header_title in enumerate(table_headers, start=left_margin):
        cell = sheet.cell(row=table_header_row, column=col_index)
        cell.value = header_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet.column_dimensions[cell.column_letter].width = 20  # Set column width

    # Populating the data
    for row_index, row_data in enumerate(all_data, start=table_header_row + 1):
        for col_index in range(1, len(table_headers) + 1):  # Skip the first column
            cell = sheet.cell(row=row_index, column=left_margin + col_index - 1)
            if col_index <= len(row_data):
                cell.value = row_data[col_index - 1]
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    # Auto-size columns after populating data
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column].width = adjusted_width

    file_path = 'temp.xlsx'
    wb.save(file_path)

    # Serve the file as a response
    response = make_response(open(file_path, 'rb').read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=Violators_Settled_Reports.xlsx'

    # Clean up (delete the temporary file)
    os.remove(file_path)

    return response